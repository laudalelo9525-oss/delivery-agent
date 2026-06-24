#!/usr/bin/env python3
"""
WEB-BASED AUTONOMOUS DELIVERY REPORT AGENT
Flask Backend Server
Multi-user, role-based, cloud-ready
"""

from flask import Flask, render_template, request, jsonify, session, send_file
from flask_socketio import SocketIO, emit, join_room, leave_room
from functools import wraps
import os
import json
import sqlite3
import threading
from datetime import datetime, timedelta
from pathlib import Path
import pytesseract
from PIL import Image
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging
from email_handler import EmailHandler

# Configuration
app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'bmp', 'tiff'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Role-based access control
ROLES = {
    'admin': ['view', 'process', 'config', 'users', 'logs'],
    'manager': ['view', 'process', 'logs'],
    'operator': ['process']
}

# User list (in production, use a real database)
USERS = {
    'mousa': {'role': 'admin', 'password': 'admin123'},
    'logistics': {'role': 'manager', 'password': 'logistics123'},
    'operator1': {'role': 'operator', 'password': 'operator123'}
}

# Global agent instance
agent = None
email_handler = None
processing_lock = threading.Lock()


class AgentCore:
    """Core agent processing engine"""
    
    def __init__(self):
        self.db_file = 'agent_database.db'
        self.config_file = 'agent_config.json'
        self.init_database()
        self.load_config()
    
    def init_database(self):
        """Initialize database"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS processed_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                user TEXT,
                image_file TEXT,
                records_count INTEGER,
                sheet_name TEXT,
                status TEXT,
                error_message TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS delivery_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_id INTEGER,
                project_number TEXT,
                trailer_no TEXT,
                trailer_type TEXT,
                element TEXT,
                count INTEGER,
                dn TEXT,
                volume REAL,
                weight REAL,
                remarks TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS agent_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                user TEXT,
                event_type TEXT,
                message TEXT,
                status TEXT
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def load_config(self):
        """Load configuration"""
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as f:
                self.config = json.load(f)
        else:
            self.config = {
                'excel_file': '',
                'email_enabled': False,
                'email_address': '',
                'email_password': '',
                'email_recipients': [],
                'daily_summary_time': '08:10'
            }
    
    def save_config(self, config_data):
        """Save configuration"""
        with open(self.config_file, 'w') as f:
            json.dump(config_data, f, indent=4)
        self.config = config_data
    
    def extract_text_from_image(self, image_path):
        """Extract text using OCR"""
        try:
            image = Image.open(image_path)
            text = pytesseract.image_to_string(image)
            return text
        except Exception as e:
            logger.error(f"OCR error: {str(e)}")
            return None
    
    def parse_delivery_data(self, text):
        """Parse OCR text into records"""
        lines = text.strip().split('\n')
        records = []
        
        for line in lines:
            if not line.strip():
                continue
            
            parts = re.split(r'\s{2,}|\t|,', line.strip())
            parts = [p.strip() for p in parts if p.strip()]
            
            if len(parts) < 6:
                continue
            
            try:
                record = {
                    'SR': len(records) + 1,
                    'Project Number': parts[0],
                    'Trailer No': parts[1],
                    'Type': parts[2],
                    'Element': parts[3],
                    'Count': int(parts[4]) if str(parts[4]).isdigit() else '',
                    'DN': parts[5],
                    'Volume': float(parts[6].replace(',', '.')) if len(parts) > 6 else '',
                    'Weight': float(parts[7].replace(',', '.')) if len(parts) > 7 else '',
                    'Remarks': parts[8] if len(parts) > 8 else ''
                }
                records.append(record)
            except (ValueError, IndexError):
                continue
        
        return records
    
    def update_excel(self, excel_file, sheet_name, records):
        """Update Excel file"""
        try:
            wb = load_workbook(excel_file)
            
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            
            ws = wb.create_sheet(sheet_name, 1)
            
            # Add headers
            ws['B1'] = 'Daily delivery tracking report'
            ws['J1'] = 'Date'
            ws['K1'] = datetime.strptime(sheet_name, '%d-%m-%Y')
            
            headers = ['', 'SR', 'Project Number', 'Project  name ', 'Location', 'Trailer No',
                      'Type', 'Element', 'Count', 'DN', ' Volume', 'Weight', 'Remarks']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add data
            for idx, record in enumerate(records, start=5):
                ws.cell(row=idx, column=2).value = record['SR']
                ws.cell(row=idx, column=3).value = record['Project Number']
                ws.cell(row=idx, column=4).value = f'=IFERROR(VLOOKUP(C{idx},\'Project list\'!$A:$B,2,0),"")'
                ws.cell(row=idx, column=5).value = f'=IFERROR(VLOOKUP(C{idx},\'Project list\'!$A:$C,3,0),"")'
                ws.cell(row=idx, column=6).value = record['Trailer No']
                ws.cell(row=idx, column=7).value = record['Type']
                ws.cell(row=idx, column=8).value = record['Element']
                ws.cell(row=idx, column=9).value = record['Count']
                ws.cell(row=idx, column=10).value = record['DN']
                ws.cell(row=idx, column=11).value = record['Volume']
                ws.cell(row=idx, column=12).value = record['Weight']
                ws.cell(row=idx, column=13).value = record['Remarks']
                
                for col_num in range(1, 14):
                    ws.cell(row=idx, column=col_num).border = thin_border
            
            wb.save(excel_file)
            return True
            
        except Exception as e:
            logger.error(f"Excel error: {str(e)}")
            return False
    
    def process_image(self, image_file, excel_file, delivery_date, user):
        """Process single image"""
        try:
            # Extract text
            text = self.extract_text_from_image(image_file)
            if not text:
                return False, []
            
            # Parse records
            records = self.parse_delivery_data(text)
            if not records:
                return False, []
            
            # Update Excel
            if self.update_excel(excel_file, delivery_date, records):
                self.record_processing(image_file, records, delivery_date, user, 'SUCCESS')
                return True, records
            else:
                self.record_processing(image_file, records, delivery_date, user, 'FAILED')
                return False, records
                
        except Exception as e:
            self.record_processing(image_file, [], delivery_date, user, 'FAILED', str(e))
            return False, []
    
    def record_processing(self, image_file, records, sheet_name, user, status, error=''):
        """Record in database"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO processed_reports (user, image_file, records_count, sheet_name, status, error_message)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (user, image_file, len(records), sheet_name, status, error))
        
        report_id = cursor.lastrowid
        
        for record in records:
            cursor.execute('''
                INSERT INTO delivery_records 
                (report_id, project_number, trailer_no, trailer_type, element, count, dn, volume, weight, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                report_id, record['Project Number'], record['Trailer No'],
                record['Type'], record['Element'], record['Count'],
                record['DN'], record['Volume'], record['Weight'], record['Remarks']
            ))
        
        conn.commit()
        conn.close()
    
    def get_daily_stats(self, date):
        """Get daily statistics"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT COUNT(*) as total_reports, SUM(records_count) as total_records
            FROM processed_reports
            WHERE DATE(timestamp) = ?
        ''', (date,))
        
        result = cursor.fetchone()
        conn.close()
        
        return {
            'total_reports': result[0] or 0,
            'total_records': result[1] or 0
        }
    
    def get_history(self, days=7):
        """Get processing history"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT timestamp, user, image_file, records_count, status
            FROM processed_reports
            WHERE timestamp >= datetime('now', '-' || ? || ' days')
            ORDER BY timestamp DESC
        ''', (days,))
        
        results = cursor.fetchall()
        conn.close()
        
        return results


# Decorators for role-based access
def require_role(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user' not in session:
                return jsonify({'error': 'Not authenticated'}), 401
            
            user_role = USERS.get(session['user'], {}).get('role')
            if user_role not in roles:
                return jsonify({'error': 'Insufficient permissions'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# Routes
@app.route('/')
def index():
    """Main dashboard"""
    return render_template('index.html')


@app.route('/api/status')
def api_status():
    """Get system status"""
    today = datetime.now().strftime('%Y-%m-%d')
    stats = agent.get_daily_stats(today)
    
    return jsonify({
        'status': 'running',
        'today': today,
        'reports': stats['total_reports'],
        'records': stats['total_records']
    })


@app.route('/api/upload', methods=['POST'])
@require_role('operator', 'manager', 'admin')
def api_upload():
    """Upload and process image"""
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files')
        delivery_date = request.form.get('date', datetime.now().strftime('%d-%m-%Y'))
        excel_file = request.form.get('excel_file')
        user = session.get('user')
        
        results = []
        total_records = 0
        
        with processing_lock:
            for file in files:
                if file and file.filename.endswith(tuple(ALLOWED_EXTENSIONS)):
                    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
                    file.save(filepath)
                    
                    success, records = agent.process_image(filepath, excel_file, delivery_date, user)
                    
                    results.append({
                        'file': file.filename,
                        'success': success,
                        'records': len(records)
                    })
                    
                    total_records += len(records)
                    
                    # Send real-time email
                    if success and email_handler and agent.config.get('email_enabled'):
                        email_handler.send_realtime_confirmation(
                            file.filename, len(records), delivery_date
                        )
                    
                    # Emit WebSocket event
                    socketio.emit('processing_update', {
                        'file': file.filename,
                        'records': len(records),
                        'status': 'success' if success else 'failed'
                    }, broadcast=True)
        
        return jsonify({
            'success': True,
            'results': results,
            'total_records': total_records
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history')
@require_role('manager', 'admin')
def api_history():
    """Get processing history"""
    history = agent.get_history(7)
    
    return jsonify({
        'history': [
            {
                'timestamp': row[0],
                'user': row[1],
                'file': row[2],
                'records': row[3],
                'status': row[4]
            }
            for row in history
        ]
    })


@app.route('/api/config', methods=['GET', 'POST'])
@require_role('admin')
def api_config():
    """Get/set configuration"""
    if request.method == 'GET':
        return jsonify(agent.config)
    
    elif request.method == 'POST':
        data = request.get_json()
        agent.save_config(data)
        
        # Reinitialize email handler
        global email_handler
        if data.get('email_enabled'):
            email_handler = EmailHandler(data)
            email_handler.schedule_daily_summary(data.get('daily_summary_time', '08:10'))
        
        return jsonify({'success': True})


@app.route('/api/users')
@require_role('admin')
def api_users():
    """Get user list"""
    users = [{
        'username': username,
        'role': info['role']
    } for username, info in USERS.items()]
    
    return jsonify({'users': users})


# WebSocket events
@socketio.on('connect')
def handle_connect():
    """Client connected"""
    logger.info(f"Client connected: {request.sid}")
    emit('response', {'data': 'Connected to agent'})


@socketio.on('disconnect')
def handle_disconnect():
    """Client disconnected"""
    logger.info(f"Client disconnected: {request.sid}")


@socketio.on('join')
def on_join(data):
    """Join room"""
    room = data.get('room')
    join_room(room)
    emit('response', {'data': f'Joined room {room}'})


if __name__ == '__main__':
    global agent, email_handler
    
    agent = AgentCore()
    
    # Configure email handler if enabled
    if agent.config.get('email_enabled'):
        email_handler = EmailHandler(agent.config)
        email_handler.schedule_daily_summary(agent.config.get('daily_summary_time', '08:10'))
    
    # Run Flask app
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)
